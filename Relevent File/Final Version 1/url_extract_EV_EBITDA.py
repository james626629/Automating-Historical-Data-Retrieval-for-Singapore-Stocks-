#!/usr/bin/env python3
"""
Enhanced Yahoo Finance scraper with EV/EBITDA data.
Scrapes historical price data AND financial metrics (EV/EBITDA) for SG tickers.
- Fetches 5 years of historical data
- Retrieves current EV/EBITDA ratio from statistics page
- Outputs Excel with price history and financial metrics

Usage:
  py url_extract_EV_EBITDA.py D05.SI O39.SI U11.SI
"""

import os
import sys
import time
import logging
from typing import List, Optional, Tuple, Dict
from urllib.parse import urlparse
from datetime import datetime, date, timedelta

import pandas as pd
from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.edge.service import Service as EdgeService
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# URL templates
HISTORY_URL_TEMPLATE = "https://sg.finance.yahoo.com/quote/{ticker}/history?period1={start_timestamp}&period2={end_timestamp}&interval=1d&frequency=1d&includeAdjustedClose=true"
STATISTICS_URL_TEMPLATE = "https://sg.finance.yahoo.com/quote/{ticker}/key-statistics"


def get_inputs_from_args_or_config() -> List[str]:
    if len(sys.argv) > 1:
        return sys.argv[1:]
    # fallback to config.json
    import json
    try:
        with open('config.json', 'r') as f:
            cfg = json.load(f)
            return cfg.get('tickers', ["D05.SI", "O39.SI", "U11.SI"])
    except Exception:
        return ["D05.SI", "O39.SI", "U11.SI"]


def build_driver() -> webdriver.Remote:
    headless = os.getenv('HEADLESS', '1') != '0'

    # Try Chrome first
    try:
        chrome_options = ChromeOptions()
        if headless:
            chrome_options.add_argument('--headless=new')
        chrome_options.add_argument('--window-size=1400,1000')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--lang=en-US')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_prefs = {"profile.default_content_setting_values.cookies": 2}
        chrome_options.add_experimental_option('prefs', chrome_prefs)
        chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])
        chrome_options.add_experimental_option('useAutomationExtension', False)

        driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)
        driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
            'source': "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        })
        driver.set_page_load_timeout(60)
        return driver
    except Exception as e:
        logger.warning(f"Chrome not available ({e}). Falling back to Edge...")

    # Fallback to Edge
    edge_options = EdgeOptions()
    if headless:
        edge_options.add_argument('headless')
    edge_options.add_argument('window-size=1400,1000')
    edge_options.add_argument('--disable-gpu')
    edge_options.add_argument('--disable-blink-features=AutomationControlled')
    driver = webdriver.Edge(service=EdgeService(EdgeChromiumDriverManager().install()), options=edge_options)
    driver.set_page_load_timeout(60)
    return driver


def accept_cookies_if_present(driver: webdriver.Remote):
    try:
        cookie_button_xpath = "//button[.//span[contains(., 'Accept all')]] | //button[contains(., 'Agree')]"
        cookie_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, cookie_button_xpath))
        )
        cookie_button.click()
        logger.info("Accepted cookies/consent banner.")
        time.sleep(1)
    except Exception:
        logger.info("No cookie banner found or could not click it.")


def fetch_ev_ebitda(driver: webdriver.Remote, ticker: str) -> Optional[float]:
    """Fetch EV/EBITDA ratio from Yahoo Finance statistics page."""
    try:
        stats_url = STATISTICS_URL_TEMPLATE.format(ticker=ticker)
        logger.info(f"Fetching EV/EBITDA for {ticker} from: {stats_url}")
        driver.get(stats_url)
        
        accept_cookies_if_present(driver)
        
        # Wait for page to load - try multiple selectors
        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'section[data-test="qsp-statistics"]'))
            )
        except:
            # Fallback: wait for any table
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.TAG_NAME, 'table'))
            )
        
        time.sleep(3)  # Let page fully render
        
        # Parse the page
        soup = BeautifulSoup(driver.page_source, 'lxml')
        
        # Debug: Save page source for inspection
        logger.debug(f"Looking for EV/EBITDA on statistics page for {ticker}")
        
        # Method 1: Look for EV/EBITDA in tables - be very specific
        tables = soup.find_all('table')
        for table_idx, table in enumerate(tables):
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all(['td', 'th'])
                if len(cells) >= 2:
                    label_text = cells[0].get_text(strip=True)
                    # Be very specific - look for exact match
                    if 'Enterprise Value/EBITDA' in label_text or 'EV/EBITDA' in label_text:
                        value_cell = cells[1]
                        value_text = value_cell.get_text(strip=True)
                        logger.info(f"Found EV/EBITDA label: '{label_text}' with value: '{value_text}'")
                        try:
                            if value_text in ['N/A', '-', '', 'NA', '--']:
                                logger.info(f"EV/EBITDA for {ticker} is not available (N/A)")
                                return None
                            # Remove any trailing text like (ttm) or (mrq)
                            value_text = value_text.split('(')[0].strip()
                            value_text = value_text.replace(',', '')
                            value = float(value_text)
                            logger.info(f"Successfully extracted EV/EBITDA for {ticker}: {value}")
                            return value
                        except ValueError:
                            logger.warning(f"Could not parse EV/EBITDA value: '{value_text}'")
        
        # Method 2: More flexible search with careful filtering
        # Look specifically in the valuation measures section
        all_rows = soup.find_all('tr')
        for row in all_rows:
            # Get all text in the row
            row_text = row.get_text(strip=True)
            # Check if this row contains EV/EBITDA (case insensitive but specific)
            if 'enterprise value/ebitda' in row_text.lower() and 'trailing' not in row_text.lower():
                cells = row.find_all(['td', 'th'])
                if len(cells) >= 2:
                    # The value should be in the second cell
                    value_text = cells[-1].get_text(strip=True)  # Use last cell in case of multiple
                    logger.info(f"Found potential EV/EBITDA row with value: '{value_text}'")
                    try:
                        if value_text in ['N/A', '-', '', 'NA', '--']:
                            return None
                        value_text = value_text.split('(')[0].strip().replace(',', '')
                        # Validate it's a reasonable number for EV/EBITDA (typically between -100 and 100)
                        value = float(value_text)
                        if -1000 < value < 1000:  # Sanity check
                            logger.info(f"Found EV/EBITDA for {ticker}: {value}")
                            return value
                    except (ValueError, AttributeError):
                        continue
        
        # Method 3: Look for specific parent containers with "Enterprise Value" heading
        # Sometimes the data is structured differently
        ev_sections = soup.find_all(text=lambda t: t and 'Enterprise Value/EBITDA' in t)
        for ev_text in ev_sections:
            parent = ev_text.parent
            # Look for the value in siblings or nearby elements
            if parent:
                siblings = parent.find_next_siblings()
                for sibling in siblings[:5]:  # Check next few siblings
                    value_text = sibling.get_text(strip=True)
                    if value_text and value_text[0].isdigit() or value_text[0] == '-':
                        try:
                            value_text = value_text.split('(')[0].strip().replace(',', '')
                            value = float(value_text)
                            if -1000 < value < 1000:
                                logger.info(f"Found EV/EBITDA for {ticker}: {value}")
                                return value
                        except ValueError:
                            continue
        
        logger.warning(f"EV/EBITDA not found for {ticker} - metric may not be available")
        return None
        
    except TimeoutException:
        logger.error(f"Page load timeout for {ticker} statistics page")
        return None
    except Exception as e:
        logger.error(f"Error fetching EV/EBITDA for {ticker}: {str(e)[:100]}")
        return None


def count_data_rows(driver: webdriver.Remote) -> int:
    try:
        rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
        valid = 0
        for r in rows:
            tds = r.find_elements(By.TAG_NAME, 'td')
            if len(tds) >= 6:  # date + OHLC + adj close + volume
                valid += 1
        return valid
    except Exception:
        return 0


def scroll_to_load_all_rows(driver: webdriver.Remote, max_loops: int = 50, patience: int = 3):
    logger.info("Scrolling to load all rows...")
    last_count = -1
    stable_rounds = 0
    for i in range(max_loops):
        current = count_data_rows(driver)
        logger.debug(f"Scroll {i+1}: {current} rows")
        if current == last_count:
            stable_rounds += 1
        else:
            stable_rounds = 0
        if stable_rounds >= patience:
            logger.info(f"Rows stabilized at {current}. Stopping scroll.")
            break
        last_count = current
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1.5)
    else:
        logger.info(f"Reached max scroll loops with {last_count} rows")


def parse_table_html(html: str) -> pd.DataFrame:
    soup = BeautifulSoup(html, 'lxml')
    table = soup.find('table', {'data-test': 'historical-prices'})
    if table is None:
        for tbl in soup.find_all('table'):
            thead = tbl.find('thead')
            if thead:
                header_text = ' '.join(thead.get_text(separator=' ').strip().lower().split())
                if all(k in header_text for k in ['date', 'open', 'high', 'low', 'close', 'volume']):
                    table = tbl
                    break
    if table is None:
        raise ValueError('Historical data table not found')

    rows = []
    for tr in table.select('tbody tr'):
        td_cells = tr.find_all('td')
        if len(td_cells) < 7:
            continue
            
        cells = [c.get_text(strip=True) for c in td_cells]
        
        # Skip Dividend / Stock Split events
        if 'Dividend' in cells[1] or 'Stock Split' in cells[1]:
            continue
        rows.append(cells)

    df = pd.DataFrame(rows, columns=['Date', 'Open', 'High', 'Low', 'Close', 'Adj Close', 'Volume'])
    if df.empty:
        return df

    # Convert types
    def to_float(x: str) -> Optional[float]:
        try:
            x = x.replace(',', '')
            if x in ('', '-', 'N/A'): return None
            return float(x)
        except Exception:
            return None

    def to_int(x: str) -> Optional[int]:
        try:
            x = x.replace(',', '')
            if x in ('', '-', 'N/A'): return None
            return int(float(x))
        except Exception:
            return None

    for col in ['Open', 'High', 'Low', 'Close', 'Adj Close']:
        df[col] = df[col].apply(to_float)
    df['Volume'] = df['Volume'].apply(to_int)

    # Convert date and remove timestamp
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    df = df.dropna(subset=['Date'])
    df['Date'] = df['Date'].dt.date
    df = df.sort_values('Date').reset_index(drop=True)

    # Daily Return = Close - Open
    df['Daily Return'] = (df['Close'] - df['Open']).round(4)

    return df


def scrape_ticker_data(driver: webdriver.Remote, ticker: str) -> Tuple[pd.DataFrame, Optional[float]]:
    """Scrape both historical data and EV/EBITDA for a ticker."""
    
    # First, get the EV/EBITDA ratio
    ev_ebitda = fetch_ev_ebitda(driver, ticker)
    
    # Then get historical data
    end_date = datetime.now()
    start_date = end_date - timedelta(days=5 * 365)
    start_timestamp = int(time.mktime(start_date.timetuple()))
    end_timestamp = int(time.mktime(end_date.timetuple()))
    history_url = HISTORY_URL_TEMPLATE.format(
        ticker=ticker,
        start_timestamp=start_timestamp,
        end_timestamp=end_timestamp
    )
    
    logger.info(f"Fetching historical data for {ticker}")
    driver.get(history_url)
    
    accept_cookies_if_present(driver)
    
    try:
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'table'))
        )
        time.sleep(2)
    except Exception:
        logger.error(f"Could not find table for {ticker}")
        return pd.DataFrame(), ev_ebitda
    
    scroll_to_load_all_rows(driver)
    
    html = driver.page_source
    df = parse_table_html(html)
    
    # Add EV/EBITDA column if we have the value
    if not df.empty and ev_ebitda is not None:
        df['EV/EBITDA'] = ev_ebitda
    
    logger.info(f"Parsed {len(df)} rows for {ticker} with EV/EBITDA: {ev_ebitda}")
    return df, ev_ebitda


def main():
    tickers = get_inputs_from_args_or_config()
    logger.info(f"Tickers to scrape: {tickers}")

    driver = build_driver()
    all_data = {}
    ev_ebitda_summary = {}
    failed = []

    try:
        for ticker in tickers:
            try:
                df, ev_ebitda = scrape_ticker_data(driver, ticker)
                if df is None or df.empty:
                    raise ValueError('No data extracted')
                all_data[ticker] = df
                ev_ebitda_summary[ticker] = ev_ebitda
                time.sleep(3)  # Be polite between requests
            except Exception as e:
                logger.error(f"Failed to scrape '{ticker}': {e}")
                failed.append(ticker)
    finally:
        driver.quit()

    if not all_data:
        logger.error("No data scraped for any ticker. Exiting.")
        sys.exit(1)

    # Create summary DataFrame
    summary_data = []
    for ticker in all_data:
        df = all_data[ticker]
        latest_close = df.iloc[0]['Close'] if not df.empty else None
        summary_data.append({
            'Ticker': ticker,
            'Latest Close': latest_close,
            'EV/EBITDA': ev_ebitda_summary.get(ticker, 'N/A'),
            'Data Points': len(df),
            'Date Range': f"{df['Date'].min()} to {df['Date'].max()}" if not df.empty else 'N/A'
        })
    
    summary_df = pd.DataFrame(summary_data)

    # Export to Excel with multiple sheets
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_file = f'sgx_stocks_with_EV_EBITDA_{timestamp}.xlsx'
    
    with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
        # Write summary sheet first
        summary_df.to_excel(writer, index=False, sheet_name='Summary')
        
        # Write individual stock sheets
        for ticker, df in all_data.items():
            sheet_name = f"{ticker}_history"[:31]
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            
            # Add formatting for EV/EBITDA column if it exists
            if 'EV/EBITDA' in df.columns:
                worksheet = writer.sheets[sheet_name]
                # Find the EV/EBITDA column
                for col_num, col in enumerate(df.columns, 1):
                    if col == 'EV/EBITDA':
                        # Add header formatting
                        from openpyxl.styles import Font, PatternFill
                        worksheet.cell(row=1, column=col_num).font = Font(bold=True)
                        worksheet.cell(row=1, column=col_num).fill = PatternFill(
                            start_color='FFFF00', end_color='FFFF00', fill_type='solid'
                        )
    
    logger.info(f"Saved Excel workbook: {out_file}")
    
    # Print summary
    print("\n=== EV/EBITDA Summary ===")
    for ticker, ev_ebitda in ev_ebitda_summary.items():
        print(f"{ticker}: {ev_ebitda if ev_ebitda else 'N/A'}")
    
    if failed:
        logger.warning(f"Tickers that failed: {failed}")


if __name__ == '__main__':
    main()
