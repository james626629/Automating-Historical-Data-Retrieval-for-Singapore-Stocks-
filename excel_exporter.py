"""
Excel Export Module for Singapore Stock Data
Handles formatting and exporting of stock data to Excel workbook
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
import logging
from datetime import datetime
import json

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Class to handle Excel export with formatting"""
    
    def __init__(self, stock_data, output_file='singapore_stocks_data.xlsx'):
        """Initialize the Excel exporter"""
        self.stock_data = stock_data
        self.output_file = output_file
        self.wb = Workbook()
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        
        # Load config for ticker names
        try:
            with open('config.json', 'r') as f:
                self.config = json.load(f)
                self.ticker_names = self.config.get('ticker_names', {})
        except:
            self.ticker_names = {}
    
    def export(self):
        """Export all stock data to Excel with formatting"""
        # Create summary sheet first
        self.create_summary_sheet()
        
        # Create individual sheets for each stock
        for ticker, data in self.stock_data.items():
            self.create_stock_sheet(ticker, data)
        
        # Save the workbook
        self.wb.save(self.output_file)
        logger.info(f"Excel file saved: {self.output_file}")
        
        return self.output_file
    
    def create_summary_sheet(self):
        """Create a summary sheet with overview of all stocks"""
        ws = self.wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "Singapore Stocks Historical Data Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = "Data Source: Yahoo Finance (yfinance)"
        ws['A4'] = "Period: Last 5 Years"
        
        # Headers for summary table
        headers = ['Ticker', 'Company Name', 'Records', 'Latest Close', 'Latest Volume', 
                   'YTD Return (%)', '1Y Return (%)', 'Max Price', 'Min Price', 'Avg Volume']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Populate summary data
        row = 7
        for ticker, data in self.stock_data.items():
            if not data.empty and 'Adjusted Close' in data.columns:
                ws.cell(row=row, column=1, value=ticker)
                ws.cell(row=row, column=2, value=self.ticker_names.get(ticker, ticker))
                ws.cell(row=row, column=3, value=len(data))
                ws.cell(row=row, column=4, value=round(data['Adjusted Close'].iloc[-1], 2))
                ws.cell(row=row, column=5, value=int(data['Volume'].iloc[-1]))
                
                # Calculate returns
                ytd_return = self.calculate_ytd_return(data)
                one_year_return = self.calculate_one_year_return(data)
                
                ws.cell(row=row, column=6, value=round(ytd_return, 2) if ytd_return else "N/A")
                ws.cell(row=row, column=7, value=round(one_year_return, 2) if one_year_return else "N/A")
                ws.cell(row=row, column=8, value=round(data['Adjusted Close'].max(), 2))
                ws.cell(row=row, column=9, value=round(data['Adjusted Close'].min(), 2))
                ws.cell(row=row, column=10, value=int(data['Volume'].mean()))
            else:
                ws.cell(row=row, column=1, value=ticker)
                ws.cell(row=row, column=2, value=self.ticker_names.get(ticker, ticker))
                ws.cell(row=row, column=3, value="No Data")
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_stock_sheet(self, ticker, data):
        """Create individual sheet for each stock"""
        # Clean sheet name (remove special characters)
        sheet_name = ticker.replace('.', '_')
        ws = self.wb.create_sheet(sheet_name)
        
        # Add title
        company_name = self.ticker_names.get(ticker, ticker)
        ws['A1'] = f"{company_name} ({ticker})"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"Historical Data - Last 5 Years"
        
        if data.empty or 'Adjusted Close' not in data.columns:
            ws['A4'] = "No data available for this ticker"
            return
        
        # Add data starting from row 4
        start_row = 4
        
        # Write headers
        for col, header in enumerate(data.columns, 1):
            cell = ws.cell(row=start_row, column=col, value=str(header))
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=False), start_row + 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Format numbers
                if c_idx > 1 and isinstance(value, (int, float)):
                    if 'Return' in str(data.columns[c_idx-1]) or 'Volatility' in str(data.columns[c_idx-1]):
                        cell.number_format = '0.00%' if value else '0.00'
                    elif 'Volume' in str(data.columns[c_idx-1]):
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = '#,##0.00'
                
                # Format dates
                if c_idx == 1:  # Date column
                    cell.number_format = 'YYYY-MM-DD'
        
        # Add chart if we have enough data
        if len(data) > 30:
            self.add_price_chart(ws, len(data) + start_row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze panes (keep headers visible)
        ws.freeze_panes = ws[f'A{start_row + 1}']
    
    def add_price_chart(self, ws, last_row):
        """Add a price chart to the worksheet"""
        chart = LineChart()
        chart.title = "Price History"
        chart.style = 2
        chart.y_axis.title = "Price"
        chart.x_axis.title = "Date"
        chart.height = 10
        chart.width = 20
        
        # Add data
        data = Reference(ws, min_col=2, min_row=4, max_col=2, max_row=last_row)
        dates = Reference(ws, min_col=1, min_row=5, max_row=last_row)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(dates)
        
        # Position the chart
        ws.add_chart(chart, "N4")
    
    def calculate_ytd_return(self, data):
        """Calculate year-to-date return"""
        try:
            current_year = datetime.now().year
            data['Date'] = pd.to_datetime(data['Date'])
            ytd_data = data[data['Date'].dt.year == current_year]
            
            if len(ytd_data) > 1:
                first_price = ytd_data['Adjusted Close'].iloc[0]
                last_price = ytd_data['Adjusted Close'].iloc[-1]
                return ((last_price - first_price) / first_price) * 100
        except:
            pass
        return None
    
    def calculate_one_year_return(self, data):
        """Calculate one-year return"""
        try:
            if len(data) >= 252:  # Approximately one trading year
                price_1y_ago = data['Adjusted Close'].iloc[-252]
                current_price = data['Adjusted Close'].iloc[-1]
                return ((current_price - price_1y_ago) / price_1y_ago) * 100
        except:
            pass
        return None
